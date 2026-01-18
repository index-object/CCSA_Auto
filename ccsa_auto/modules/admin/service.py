import logging
from typing import List, Dict, Any, Optional

from ccsa_auto.core.database import SessionLocal
from ccsa_auto.core.models import User, Task, Announcement
from ccsa_auto.modules.announcement.service import AnnouncementService
from ccsa_auto.modules.logging.service import LoggingService
from ccsa_auto.modules.task.scheduler import remove_task_from_scheduler
from ccsa_auto.utils.timezone import format_datetime_for_display, get_current_time

logger = logging.getLogger(__name__)


class AdminService:
    """管理服务"""

    @staticmethod
    def is_admin(user_id):
        """检查是否为管理员"""
        db = SessionLocal()
        try:
            user = db.query(User).filter_by(id=int(user_id)).first()
            return user and user.is_admin
        finally:
            db.close()

    @staticmethod
    def get_users(keyword: str = None, status: int = None) -> Dict[str, Any]:
        """获取用户列表"""
        db = SessionLocal()
        try:
            query = db.query(User)
            if keyword:
                query = query.filter(
                    (User.username.contains(keyword))
                    | (User.company_name.contains(keyword))
                )
            if status is not None:
                query = query.filter_by(status=status)

            users = query.order_by(User.created_at.desc()).all()

            user_list = []
            for user in users:
                user_list.append(
                    {
                        "id": user.id,
                        "username": user.username,
                        "name": user.name,
                        "external_username": user.external_username,
                        "company_name": user.company_name,
                        "status": user.status,
                        "is_admin": user.is_admin,
                        "created_at": format_datetime_for_display(user.created_at),
                    }
                )

            return {"success": True, "users": user_list}
        finally:
            db.close()

    @staticmethod
    def update_user_status(
        user_id: int, status: int, admin_id: int = None
    ) -> Dict[str, Any]:
        """修改用户状态，同时联动处理定时任务"""
        db = SessionLocal()
        try:
            user = db.query(User).filter_by(id=user_id).first()

            if not user:
                return {"success": False, "message": "用户不存在"}

            if user.username == "admin":
                return {"success": False, "message": "不能修改管理员状态"}

            old_status = user.status
            user.status = status
            db.commit()

            if old_status == status:
                return {"success": True, "message": "状态未变化"}

            if old_status == 0 and status == 1:
                AdminService._sync_tasks_on_ban(db, user_id, ban=True)
                content = f"封禁用户 {user.username}"
            elif old_status == 1 and status == 0:
                AdminService._sync_tasks_on_ban(db, user_id, ban=False)
                content = f"解封用户 {user.username}"
            else:
                content = f"修改用户 {user.username} 状态: {old_status} → {status}"

            if admin_id:
                LoggingService.log_admin_operation(
                    admin_id=admin_id,
                    operation="UPDATE_USER_STATUS",
                    target_type="user",
                    target_id=user_id,
                    content=content,
                )

            return {"success": True, "message": "用户状态更新成功"}
        except Exception as e:
            db.rollback()
            logger.exception("更新用户状态失败")
            return {"success": False, "message": f"更新用户状态失败: {str(e)}"}
        finally:
            db.close()

    @staticmethod
    def batch_update_user_status(
        user_ids: List[int], status: int, admin_id: int = None
    ) -> Dict[str, Any]:
        """批量修改用户状态"""
        db = SessionLocal()
        try:
            results = []
            success_count = 0
            fail_count = 0

            for user_id in user_ids:
                user = db.query(User).filter_by(id=user_id).first()
                if not user:
                    results.append(
                        {"user_id": user_id, "success": False, "message": "用户不存在"}
                    )
                    fail_count += 1
                    continue
                if user.username == "admin":
                    results.append(
                        {
                            "user_id": user_id,
                            "success": False,
                            "message": "不能修改管理员",
                        }
                    )
                    fail_count += 1
                    continue

                old_status = user.status
                user.status = status
                db.commit()

                if old_status == 0 and status == 1:
                    AdminService._sync_tasks_on_ban(db, user_id, ban=True)
                elif old_status == 1 and status == 0:
                    AdminService._sync_tasks_on_ban(db, user_id, ban=False)

                results.append(
                    {"user_id": user_id, "success": True, "message": "操作成功"}
                )
                success_count += 1

            if admin_id:
                LoggingService.log_admin_operation(
                    admin_id=admin_id,
                    operation="BATCH_UPDATE_USER_STATUS",
                    target_type="user",
                    target_id=None,
                    content=f"批量{('封禁' if status == 1 else '解封')}用户 {success_count} 个",
                )

            return {
                "success": True,
                "results": results,
                "message": f"成功: {success_count}, 失败: {fail_count}",
            }
        except Exception as e:
            db.rollback()
            logger.exception("批量更新用户状态失败")
            return {"success": False, "message": f"批量更新用户状态失败: {str(e)}"}
        finally:
            db.close()

    @staticmethod
    def _sync_tasks_on_ban(db, user_id: int, ban: bool):
        """封禁/解封用户时联动处理任务"""
        from ccsa_auto.modules.task.scheduler import (
            add_task_to_scheduler,
            remove_task_from_scheduler,
        )
        from ccsa_auto.utils.timezone import calculate_next_run_time_utc

        tasks = db.query(Task).filter_by(user_id=user_id).all()
        current_time = get_current_time()

        for task in tasks:
            if ban:
                task.is_active = False
                remove_task_from_scheduler(task.id)
            else:
                task.is_active = True
                new_run_time = calculate_next_run_time_utc(task, current_time)
                task.next_run_time = new_run_time
                add_task_to_scheduler(task.id)

        db.commit()
        logger.info(
            f"用户 {user_id} 状态变更，{'禁用' if ban else '恢复'}了 {len(tasks)} 个任务"
        )

    @staticmethod
    def delete_user(user_id: int, admin_id: int = None) -> Dict[str, Any]:
        """删除用户"""
        db = SessionLocal()
        try:
            user = db.query(User).filter_by(id=user_id).first()

            if not user:
                return {"success": False, "message": "用户不存在"}

            if user.username == "admin":
                return {"success": False, "message": "不能删除管理员"}

            username = user.username

            for task in user.tasks:
                remove_task_from_scheduler(task.id)

            db.query(Task).filter_by(user_id=user_id).delete()
            db.delete(user)
            db.commit()

            if admin_id:
                LoggingService.log_admin_operation(
                    admin_id=admin_id,
                    operation="DELETE_USER",
                    target_type="user",
                    target_id=user_id,
                    content=f"删除用户 {username}",
                )

            return {"success": True, "message": "用户删除成功"}
        except Exception as e:
            db.rollback()
            logger.exception("删除用户失败")
            return {"success": False, "message": f"删除用户失败: {str(e)}"}
        finally:
            db.close()

    @staticmethod
    def get_all_tasks(
        keyword: str = None, task_type: str = None, execution_status: str = None
    ) -> Dict[str, Any]:
        """获取所有任务"""
        db = SessionLocal()
        try:
            query = db.query(Task)

            if task_type:
                query = query.filter_by(task_type=task_type)
            if execution_status:
                query = query.filter_by(execution_status=execution_status)

            tasks = query.order_by(Task.created_at.desc()).all()

            task_list = []
            for task in tasks:
                user = db.query(User).filter_by(id=task.user_id).first()
                if keyword:
                    if user:
                        if keyword not in user.username and keyword not in (
                            user.company_name or ""
                        ):
                            continue
                    else:
                        continue

                task_list.append(
                    {
                        "id": task.id,
                        "user_id": task.user_id,
                        "username": user.username if user else "",
                        "task_type": task.task_type,
                        "task_name": task.task_name,
                        "is_active": task.is_active,
                        "execution_status": task.execution_status,
                        "external_status": task.external_status,
                        "result": task.result,
                        "scheduled_time": format_datetime_for_display(
                            task.next_run_time
                        ),
                        "executed_at": format_datetime_for_display(task.executed_at),
                        "created_at": format_datetime_for_display(task.created_at),
                    }
                )

            return {"success": True, "tasks": task_list}
        finally:
            db.close()

    @staticmethod
    def toggle_task(
        task_id: int, is_active: bool, admin_id: int = None
    ) -> Dict[str, Any]:
        """启用/禁用任务"""
        from ccsa_auto.modules.task.scheduler import (
            add_task_to_scheduler,
            remove_task_from_scheduler,
        )
        from ccsa_auto.utils.timezone import (
            calculate_next_run_time_utc,
            get_current_time,
        )

        db = SessionLocal()
        try:
            task = db.query(Task).filter_by(id=task_id).first()
            if not task:
                return {"success": False, "message": "任务不存在"}

            task.is_active = is_active

            if is_active:
                current_time = get_current_time()
                new_run_time = calculate_next_run_time_utc(task, current_time)
                task.next_run_time = new_run_time
                add_task_to_scheduler(task.id)
                message = "任务已启用"
            else:
                remove_task_from_scheduler(task.id)
                message = "任务已禁用"

            db.commit()

            if admin_id:
                LoggingService.log_admin_operation(
                    admin_id=admin_id,
                    operation="TOGGLE_TASK",
                    target_type="task",
                    target_id=task_id,
                    content=f"{'启用' if is_active else '禁用'}任务 {task.task_name}",
                )

            return {"success": True, "message": message}
        except Exception as e:
            db.rollback()
            logger.exception("切换任务状态失败")
            return {"success": False, "message": f"切换任务状态失败: {str(e)}"}
        finally:
            db.close()

    @staticmethod
    def trigger_task(task_id: int, admin_id: int = None) -> Dict[str, Any]:
        """手动触发任务执行"""
        from ccsa_auto.modules.task.scheduler import execute_user_task

        db = SessionLocal()
        try:
            task = db.query(Task).filter_by(id=task_id).first()
            if not task:
                return {"success": False, "message": "任务不存在"}

            if not task.is_active:
                return {"success": False, "message": "任务已禁用，无法执行"}

            if admin_id:
                LoggingService.log_admin_operation(
                    admin_id=admin_id,
                    operation="TRIGGER_TASK",
                    target_type="task",
                    target_id=task_id,
                    content=f"手动触发任务 {task.task_name}",
                )

            execute_user_task(task_id)

            return {"success": True, "message": "任务已触发执行"}
        except Exception as e:
            logger.exception("触发任务失败")
            return {"success": False, "message": f"触发任务失败: {str(e)}"}
        finally:
            db.close()

    @staticmethod
    def delete_task(task_id: int, admin_id: int = None) -> Dict[str, Any]:
        """删除任务"""
        from ccsa_auto.modules.task.scheduler import remove_task_from_scheduler

        db = SessionLocal()
        try:
            task = db.query(Task).filter_by(id=task_id).first()
            if not task:
                return {"success": False, "message": "任务不存在"}

            task_name = task.task_name
            remove_task_from_scheduler(task.id)
            db.delete(task)
            db.commit()

            if admin_id:
                LoggingService.log_admin_operation(
                    admin_id=admin_id,
                    operation="DELETE_TASK",
                    target_type="task",
                    target_id=task_id,
                    content=f"删除任务 {task_name}",
                )

            return {"success": True, "message": "任务删除成功"}
        except Exception as e:
            db.rollback()
            logger.exception("删除任务失败")
            return {"success": False, "message": f"删除任务失败: {str(e)}"}
        finally:
            db.close()

    @staticmethod
    def get_statistics() -> Dict[str, Any]:
        """获取平台统计数据"""
        from datetime import datetime, timedelta
        from ccsa_auto.utils.timezone import get_current_time

        db = SessionLocal()
        try:
            now = get_current_time()
            week_ago = now - timedelta(days=7)
            month_ago = now - timedelta(days=30)

            total_users = db.query(User).count()
            active_users = db.query(User).filter_by(status=0).count()
            banned_users = db.query(User).filter_by(status=1).count()

            new_users_week = db.query(User).filter(User.created_at >= week_ago).count()
            new_users_month = (
                db.query(User).filter(User.created_at >= month_ago).count()
            )

            total_tasks = db.query(Task).count()
            active_tasks = db.query(Task).filter_by(is_active=True).count()
            pending_tasks = db.query(Task).filter_by(execution_status="pending").count()
            completed_tasks = (
                db.query(Task).filter_by(execution_status="completed").count()
            )
            failed_tasks = db.query(Task).filter_by(execution_status="failed").count()
            running_tasks = db.query(Task).filter_by(execution_status="running").count()

            completed_week = (
                db.query(Task)
                .filter(
                    Task.execution_status == "completed", Task.executed_at >= week_ago
                )
                .count()
            )
            completed_month = (
                db.query(Task)
                .filter(
                    Task.execution_status == "completed", Task.executed_at >= month_ago
                )
                .count()
            )

            total_announcements = db.query(Announcement).count()

            return {
                "success": True,
                "data": {
                    "users": {
                        "total": total_users,
                        "active": active_users,
                        "banned": banned_users,
                        "new_week": new_users_week,
                        "new_month": new_users_month,
                    },
                    "tasks": {
                        "total": total_tasks,
                        "active": active_tasks,
                        "pending": pending_tasks,
                        "completed": completed_tasks,
                        "failed": failed_tasks,
                        "running": running_tasks,
                        "completed_week": completed_week,
                        "completed_month": completed_month,
                    },
                    "announcements": {"total": total_announcements},
                },
            }
        finally:
            db.close()

    @staticmethod
    def get_daily_stats(days: int = 7) -> Dict[str, Any]:
        """获取每日统计趋势"""
        from datetime import datetime, timedelta
        from ccsa_auto.utils.timezone import get_current_time, SHANGHAI_TZ

        db = SessionLocal()
        try:
            now = get_current_time()
            stats = []

            for i in range(days):
                date = now - timedelta(days=i)
                date_start = date.replace(hour=0, minute=0, second=0, microsecond=0)
                date_end = date_start + timedelta(days=1)

                user_count = (
                    db.query(User)
                    .filter(User.created_at >= date_start, User.created_at < date_end)
                    .count()
                )

                task_count = (
                    db.query(Task)
                    .filter(
                        Task.execution_status == "completed",
                        Task.executed_at >= date_start,
                        Task.executed_at < date_end,
                    )
                    .count()
                )

                stats.append(
                    {
                        "date": date.strftime("%Y-%m-%d"),
                        "new_users": user_count,
                        "completed_tasks": task_count,
                    }
                )

            return {"success": True, "stats": list(reversed(stats))}
        finally:
            db.close()

    @staticmethod
    def export_statistics_report() -> Optional[str]:
        """导出统计报表"""
        try:
            from openpyxl import Workbook
            from datetime import datetime

            wb = Workbook()

            ws_users = wb.active
            ws_users.title = "用户统计"

            result = AdminService.get_statistics()
            if not result["success"]:
                return None

            data = result["data"]

            ws_users.append(["用户统计"])
            ws_users.append(["总用户", data["users"]["total"]])
            ws_users.append(["活跃用户", data["users"]["active"]])
            ws_users.append(["封号用户", data["users"]["banned"]])
            ws_users.append(["本周新增", data["users"]["new_week"]])
            ws_users.append(["本月新增", data["users"]["new_month"]])

            ws_tasks = wb.create_sheet("任务统计")
            ws_tasks.append(["任务统计"])
            ws_tasks.append(["总任务", data["tasks"]["total"]])
            ws_tasks.append(["激活任务", data["tasks"]["active"]])
            ws_tasks.append(["待执行", data["tasks"]["pending"]])
            ws_tasks.append(["已完成", data["tasks"]["completed"]])
            ws_tasks.append(["失败", data["tasks"]["failed"]])
            ws_tasks.append(["本周完成", data["tasks"]["completed_week"]])
            ws_tasks.append(["本月完成", data["tasks"]["completed_month"]])

            ws_trend = wb.create_sheet("每日趋势")
            ws_trend.append(["日期", "新增用户", "完成任务"])

            trend_result = AdminService.get_daily_stats(30)
            if trend_result["success"]:
                for row in trend_result["stats"]:
                    ws_trend.append(
                        [row["date"], row["new_users"], row["completed_tasks"]]
                    )

            import os

            export_dir = "exports"
            if not os.path.exists(export_dir):
                os.makedirs(export_dir)

            filename = f"statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(export_dir, filename)
            wb.save(filepath)

            return filepath
        except Exception as e:
            logger.error(f"导出统计报表失败: {e}")
            return None

    @staticmethod
    def get_system_config() -> Dict[str, Any]:
        """获取系统配置"""
        from ccsa_auto.core.config import Config

        return {
            "success": True,
            "config": {
                "task_schedule": Config.TASK_SCHEDULE,
                "task_details": Config.TASK_DETAILS,
                "task_fixer_enabled": Config.TASK_FIXER_ENABLED,
                "task_fixer_cron": Config.TASK_FIXER_CRON,
                "session_timeout": Config.SESSION_TIMEOUT,
                "session_absolute_timeout": Config.SESSION_ABSOLUTE_TIMEOUT,
            },
        }

    @staticmethod
    def update_task_schedule_config(
        daily_hour_start: int,
        daily_hour_end: int,
        weekly_hour_start: int,
        weekly_hour_end: int,
        monthly_hour_start: int,
        monthly_hour_end: int,
        admin_id: int = None,
    ) -> Dict[str, Any]:
        """更新任务调度时间配置"""
        from ccsa_auto.core.config import Config

        try:
            Config.TASK_DETAILS["DAILY"]["hour_range"] = (
                daily_hour_start,
                daily_hour_end,
            )
            Config.TASK_DETAILS["WEEKLY"]["hour_range"] = (
                weekly_hour_start,
                weekly_hour_end,
            )
            Config.TASK_DETAILS["MONTHLY"]["hour_range"] = (
                monthly_hour_start,
                monthly_hour_end,
            )

            if admin_id:
                LoggingService.log_admin_operation(
                    admin_id=admin_id,
                    operation="UPDATE_TASK_SCHEDULE",
                    target_type="config",
                    target_id=None,
                    content=f"更新任务调度时间配置",
                )

            return {"success": True, "message": "任务调度时间配置已更新"}
        except Exception as e:
            logger.exception("更新任务调度配置失败")
            return {"success": False, "message": f"更新配置失败: {str(e)}"}

    @staticmethod
    def toggle_task_fixer(enabled: bool, admin_id: int = None) -> Dict[str, Any]:
        """启用/禁用任务修复器"""
        from ccsa_auto.core.config import Config
        from ccsa_auto.modules.task.scheduler import (
            toggle_task_fixer as scheduler_toggle,
        )

        try:
            Config.TASK_FIXER_ENABLED = enabled
            scheduler_toggle(enabled)

            if admin_id:
                LoggingService.log_admin_operation(
                    admin_id=admin_id,
                    operation="TOGGLE_TASK_FIXER",
                    target_type="config",
                    target_id=None,
                    content=f"{'启用' if enabled else '禁用'}任务修复器",
                )

            return {
                "success": True,
                "message": f"任务修复器已{'启用' if enabled else '禁用'}",
            }
        except Exception as e:
            logger.exception("切换任务修复器状态失败")
            return {"success": False, "message": f"操作失败: {str(e)}"}
