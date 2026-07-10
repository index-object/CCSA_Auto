import logging
from datetime import datetime, timedelta
from typing import Optional, Dict, Any, List

from ccsa_auto.core.database import SessionLocal
from ccsa_auto.modules.logging.models import AppLog

logger = logging.getLogger(__name__)


class LogType:
    OPERATION = "operation"
    TASK = "task"
    AUTH = "auth"
    SYSTEM = "system"
    ERROR = "error"


class LoggingService:
    """统一日志服务"""

    @staticmethod
    def log(
        log_type: str,
        operation: str,
        content: str,
        user_id: Optional[int] = None,
        target_type: Optional[str] = None,
        target_id: Optional[int] = None,
        ip_address: Optional[str] = None,
        status: str = "success",
    ) -> None:
        """记录日志"""
        db = SessionLocal()
        try:
            log = AppLog(
                log_type=log_type,
                operation=operation,
                content=content,
                user_id=user_id,
                target_type=target_type,
                target_id=target_id,
                ip_address=ip_address,
                status=status,
            )
            db.add(log)
            db.commit()
        except Exception as e:
            logger.error(f"记录日志失败: {e}")
        finally:
            db.close()

    @staticmethod
    def log_admin_operation(
        admin_id: int,
        operation: str,
        target_type: str,
        target_id: Optional[int],
        content: str,
        ip_address: Optional[str] = None,
    ) -> None:
        """记录管理员操作"""
        LoggingService.log(
            log_type=LogType.OPERATION,
            operation=operation,
            content=content,
            user_id=admin_id,
            target_type=target_type,
            target_id=target_id,
            ip_address=ip_address,
        )

    @staticmethod
    def log_task_execution(
        task_id: int,
        user_id: int,
        task_type: str,
        status: str,
        message: str,
    ) -> None:
        """记录任务执行"""
        LoggingService.log(
            log_type=LogType.TASK,
            operation=task_type,
            content=message,
            user_id=user_id,
            target_type="task",
            target_id=task_id,
            status=status,
        )

    @staticmethod
    def log_auth(
        user_id: Optional[int],
        action: str,
        success: bool,
        ip_address: Optional[str] = None,
        detail: Optional[str] = None,
    ) -> None:
        """记录认证事件"""
        LoggingService.log(
            log_type=LogType.AUTH,
            operation=action,
            content=detail or f"{action} {'成功' if success else '失败'}",
            user_id=user_id,
            ip_address=ip_address,
            status="success" if success else "failed",
        )

    @staticmethod
    def log_error(
        operation: str,
        content: str,
        user_id: Optional[int] = None,
        ip_address: Optional[str] = None,
    ) -> None:
        """记录错误日志"""
        LoggingService.log(
            log_type=LogType.ERROR,
            operation=operation,
            content=content,
            user_id=user_id,
            ip_address=ip_address,
            status="failed",
        )

    @staticmethod
    def get_logs(
        log_type: Optional[str] = None,
        user_id: Optional[int] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        page: int = 1,
        page_size: int = 50,
    ) -> Dict[str, Any]:
        """查询日志"""
        db = SessionLocal()
        try:
            query = db.query(AppLog)

            if log_type and log_type != "all":
                query = query.filter_by(log_type=log_type)
            if user_id is not None:
                query = query.filter_by(user_id=user_id)
            if start_date:
                query = query.filter(AppLog.created_at >= start_date)
            if end_date:
                query = query.filter(AppLog.created_at <= end_date)

            total = query.count()
            logs = (
                query.order_by(AppLog.created_at.desc())
                .offset((page - 1) * page_size)
                .limit(page_size)
                .all()
            )

            return {
                "success": True,
                "logs": [log.to_dict() for log in logs],
                "total": total,
                "page": page,
                "page_size": page_size,
            }
        except Exception as e:
            logger.error(f"查询日志失败: {e}")
            return {"success": False, "message": str(e), "logs": [], "total": 0}
        finally:
            db.close()

    @staticmethod
    def cleanup_old_logs(days: int = 60) -> int:
        """清理过期日志"""
        db = SessionLocal()
        try:
            cutoff_date = datetime.utcnow() - timedelta(days=days)
            deleted = db.query(AppLog).filter(AppLog.created_at < cutoff_date).delete()
            db.commit()
            logger.info(f"已清理 {deleted} 条 {days} 天前的日志")
            return deleted
        except Exception as e:
            db.rollback()
            logger.error(f"清理日志失败: {e}")
            return 0
        finally:
            db.close()

    @staticmethod
    def export_to_xlsx(
        log_type: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
    ) -> Optional[str]:
        """导出日志为 xlsx 文件，返回文件路径"""
        try:
            from openpyxl import Workbook

            result = LoggingService.get_logs(
                log_type=log_type,
                start_date=start_date,
                end_date=end_date,
                page=1,
                page_size=10000,
            )

            if not result.get("success"):
                return None

            wb = Workbook()
            ws = wb.active
            ws.title = "操作日志"

            headers = [
                "ID",
                "时间",
                "类型",
                "操作",
                "内容",
                "用户ID",
                "目标类型",
                "目标ID",
                "IP",
                "状态",
            ]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            for row, log in enumerate(result["logs"], 2):
                ws.cell(row=row, column=1, value=log["id"])
                ws.cell(row=row, column=2, value=log["created_at"])
                ws.cell(row=row, column=3, value=log["log_type"])
                ws.cell(row=row, column=4, value=log["operation"])
                ws.cell(row=row, column=5, value=log["content"])
                ws.cell(row=row, column=6, value=log["user_id"])
                ws.cell(row=row, column=7, value=log["target_type"])
                ws.cell(row=row, column=8, value=log["target_id"])
                ws.cell(row=row, column=9, value=log["ip_address"])
                ws.cell(row=row, column=10, value=log["status"])

            for column in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(
                    max_length + 2, 50
                )

            import os

            export_dir = "exports"
            if not os.path.exists(export_dir):
                os.makedirs(export_dir)

            filename = f"logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(export_dir, filename)
            wb.save(filepath)

            logger.info(f"日志导出成功: {filepath}")
            return filepath

        except ImportError:
            logger.error("openpyxl 未安装，无法导出 xlsx")
            return None
        except Exception as e:
            logger.error(f"导出日志失败: {e}")
            return None
